from collections import OrderedDict

import openpyxl
from openpyxl import load_workbook

import tkinter as tk
from tkinter import filedialog, messagebox
import win32com.client as win32
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def check_file_open():
    """ Функция для проверки, открыт ли файл Excel """

    messagebox.showinfo('Внимание', 'Пожалуйста, закройте таблицу Excel')


def get_data_and_headers(_file_path, _sheet=None):
    """ Получить двумерный массив данных и шапку из excel таблицы """

    workbook = openpyxl.load_workbook(_file_path)

    # Загружаем файл
    if _sheet is None:
        # Получаем первую страницу
        wb_sheet: Worksheet = workbook.active
    elif isinstance(_sheet, int):
        # Получение листа по индексу
        wb_sheet = workbook.worksheets[_sheet]  # Получение первого листа
    elif isinstance(_sheet, str):
        # Получение листа по названию
        wb_sheet = workbook[_sheet]  # Получение листа с названием "Sheet1"

    # Считать данные
    _shapka = []
    _array = []

    for x, row in enumerate(wb_sheet.iter_rows(values_only=True)):
        _vlas = []
        for cell_value in row:
            _vlas.append(cell_value)

        if x == 0:  # Пропустить шапку
            # Привести шапку к нижнему регистру
            temp = []
            for val in row:
                temp.append(str(val).lower())
            _shapka.append(temp)
            _array.append(temp)
        else:

            temp_row = []
            for s in _vlas:
                if isinstance(s, type(None)):
                    s = ""
                temp_row.append(s)
            _array.append(temp_row)

    # Закрываем файл
    workbook.close()
    return _shapka, _array


def clear_cells_according_to_the_template(_array):
    """ Очистить ячейки по шаблону """

    shapka = array[0]
    main_columns_pattern = "осн"
    main_columns = [str(col).lower() for col in shapka if main_columns_pattern in str(col).lower()]

    # Пройти по основным столбцам, проверить что в них есть вхождения слов/параметров "Рассчитывать ночное время (нет)",
    # если есть то в данной строке в столбце с вхождением "перераб" удалить значение в ячейке

    # для удаления ячеек в строке в которой есть в основных значениях - ночьные часы в стобцах "перераб"...
    need_del = ["Переработка в ночное время(Как рабочее время)", "Переработка в ночное время(Как переработка)",
                "Переработка в ночное время(Не учитывать)"]

    new_need_del_artem = [
        "Общая продолжительность мягких прогулов(больше прогула)",
        "Общая продолжительность мягких прогулов(меньше прогула)",
        "Дополнительные перерывы мягкие прогулы (Включать в рабочее время)",
        "Дополнительные перерывы мягкие прогулы (Не рассчитывать)",
        "Дополнительные перерывы мягкие прогулы (Рассчитывать как перерыв)",
        "Максимальное время перерыва(больше общей продолжительности)",
        "Максимальное время перерыва(меньше общей продолжительности)"
    ]
    new_need_del_artem = [val.lower() for val in new_need_del_artem]

    for itr, row in enumerate(_array):
        for head_val in main_columns:
            index_head_val: int
            for jj, h in enumerate(shapka):
                if h.lower() == head_val:
                    index_head_val = jj

            if "ночные часы (нет)" == str(row[index_head_val]).lower():
                for col_iter, column_name in enumerate(_array[0]):  # Пройтись по всей шапке
                    if "перераб" in str(column_name).lower():  # Найти в шапке столбец с вхождением названия "перераб
                        if _array[itr][col_iter] in need_del:
                            _array[itr][col_iter] = ""  # для строки указать по индексу столбца None

            if "свободный график (да)" == str(row[index_head_val]).lower():
                # в одной строке ->
                # Если встречается "свободный график (да)" то удалить

                for art, val in enumerate(row[: len(main_columns)]):
                    if val.lower() in new_need_del_artem:
                        _array[itr][art] = ""

                for col_iter, column_name in enumerate(_array[0]):
                    if "наруш" in str(column_name).lower():
                        _array[itr][col_iter] = ""
                    if "перераб" in str(column_name).lower():
                        _array[itr][col_iter] = ""

    # если есть 2 одинаковых параметра, то оставить 1
    for i, row in enumerate(_array):
        # по строке
        new_row = []

        uniq_strings = set()  # неповторяющееся множество
        for string in row:
            # по значениям строки
            if string not in uniq_strings:  # если есть уникальный параметр со значением, то добавить его в уники
                uniq_strings.add(string)

        for string in row:
            if string in uniq_strings:  # если значение в униках,
                new_row.append(string)  # то добавляем ее в новую строку
                uniq_strings.remove(string)  # и убираем ее из уников
            else:
                new_row.append("")  # если значение не
        _array[i] = new_row  # обновляем строку с очищенными повторами в порядке очистки с лева на право

    return _array


def duplicate_removal(_array):
    """ Удалить дубликаты """

    _shapka = _array[0]

    # Создание упорядоченного словаря для сохранения порядка значений
    unique_dict = OrderedDict()

    for row in _array[1:]:
        # Сортировка значений каждой строки от а до я без слова "ПОКРАСКА"
        sorted_row = sorted(row[:-1])

        # Преобразование строки в кортеж и использование кортежа в качестве ключа словаря
        key = tuple(sorted_row)

        # Проверка, содержит ли строка слово "ПОКРАСКА" в конце
        if row[-1] == "ПОКРАСКА":
            # Проверка, существует ли уже строка с таким ключом в словаре
            if key not in unique_dict:
                unique_dict[key] = row
        else:
            # Замена значения в словаре новой строкой
            unique_dict[key] = row

    # Сборка строк в изначальном порядке
    _array = [_shapka] + list(unique_dict.values())

    return _array


def split_a_string_with_duplicate_parameters_but_different_values(_array):
    """
    Разбить строку с дубликатами параметров, но разными значениями

    Найти одинаковые параметры и если у них разные значения то создать дубль строки в которых будут
    в единственном экземпляре повторяющиеся параметры
    """
    new_array = []

    for j, row in enumerate(_array):
        if j == 0:
            # Добавляем шапку в массив
            new_array.append(row)
            continue

        parameters = {}  # Создаем словарь для хранения параметров и их значения
        # Проверяем, есть ли один и тот же параметр с разными значениями
        same_parameter_with_different_values = False  # один и тот же параметр с разными значениями

        for item in row:
            # Получаем параметр и значение
            if item == "":
                continue
            try:
                parameter = item.split('(')[0].strip()
            except:
                parameter = ""
            try:
                value = item.split('(')[1].strip(')')
            except:
                value = ""

            # Если параметр уже есть в словаре и значение отличается, устанавливаем флаг has_duplicate в True
            if parameter in parameters and parameters[parameter] != value:
                same_parameter_with_different_values = True
                break
            # Добавляем параметр и значение в словарь
            parameters[parameter] = value

        # Если есть один и тот же параметр с разными значениями, добавляем строку в новый массив с желтым цветом
        if same_parameter_with_different_values:
            # Окрашивание ячеек в текущей строке
            temp_row1 = []
            duble_param = f"{parameter} ({value})"
            duble_param_without_a_space = f"{parameter}({value})"
            double_param_index_in_row: int
            # строка 1
            for i, item in enumerate(row):
                # Присвоить текущей строке Параметр со значением, а у повторяющегося параметра сделать пусто
                # Создать еще одну строку с параметром который стал пустым в строке выше а тот параметр сделать пустым
                if duble_param == item or duble_param_without_a_space == item:
                    item = ""
                    temp_row1.append(item)
                elif item == 'ПОКРАСКА':
                    continue
                else:
                    temp_row1.append(item)

            temp_row2 = []
            duble_param = f"{parameter} ({parameters[parameter]})"

            for i, item in enumerate(row):
                if duble_param == item or duble_param_without_a_space == item:
                    item = ""
                    temp_row2.append(item)
                elif item == 'ПОКРАСКА':
                    continue
                else:
                    temp_row2.append(item)
            temp_row1.append('ПОКРАСКА')
            temp_row2.append('ПОКРАСКА')
            new_array.append(temp_row1)
            new_array.append(temp_row2)
        else:
            new_array.append(row)

    return new_array


def save_table_in_new_sheet(_file_path, table_array, new_sheet):
    """ Сохранить таблицу в новый лист """

    # Загрузка существующей книги
    _workbook = load_workbook(_file_path)

    try:
        _workbook = load_workbook(_file_path)
        # Проверка, существует ли лист с заданным именем
        if new_sheet in _workbook.sheetnames:
            # Удаление существующего листа
            _workbook.remove(_workbook[new_sheet])
    except Exception:
        pass

    # Создание нового листа
    new_sheet = _workbook.create_sheet(title=new_sheet)

    cell_color = "FFFF00"

    for row_index, row in enumerate(table_array):
        if "ПОКРАСКА" in row:
            for col_index, val in enumerate(row):
                if val == "ПОКРАСКА":
                    pass
                else:
                    cell = new_sheet.cell(row=row_index + 1, column=col_index + 1)
                    cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
                    cell.value = val
        else:
            new_sheet.append(row)

    # Сохранение книги с новым листом
    _workbook.save(_file_path)


def open_excel_and_sheet(_file_path, sheet_name):
    # Запуск приложения Excel
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = True
    # Открытие файла
    workbook = excel.Workbooks.Open(_file_path)
    # Переход на лист с именем "New Sheet"
    sheet = workbook.Sheets(sheet_name)
    sheet.Activate()


def for_column_groups_shift_values_to_empty_cells(_array):
    """ Для групп столбцов сдвинуть значения в пустые ячейки """

    headers_patterns = ["осн", "наруш", "перераб", "проч"]
    _shapka = array[0]
    # Заполнить группы
    groups_list = []

    for header_p in headers_patterns:
        group = []
        for h_s in _shapka:
            if header_p in h_s.lower():
                group.append(h_s)
        if len(group) > 0:
            groups_list.append(group)

    for group in groups_list:  # Пройти по всем группам
        for povtor in range(len(group)):
            for group_index in range(0, (len(group) - 1)):  # Пройти по группе
                index_array_header_1: int = 0
                index_array_header_2: int = 0

                for l, h_s in enumerate(_shapka):
                    if h_s == group[group_index]:
                        index_array_header_1 = l
                    if h_s == group[group_index + 1]:
                        index_array_header_2 = l

                for i, row in enumerate(_array):  # По строкам
                    if i == 0:  # Если шапка, то пропустить
                        continue
                    if _array[0][index_array_header_1] in group and _array[0][index_array_header_2] in group:
                        # если в текущей итерации и в следующей мы находимся в группе, то пройти по всем строкам,
                        # сделать проверку на пустую ячейку в следующей строке и если есть такая ячейка, то сдвинуть
                        # ее на ту ячейку в итерации которой ы находимся
                        if row[index_array_header_1] == "" and row[index_array_header_2] != "":
                            _array[i][index_array_header_1] = _array[i][index_array_header_2]
                            _array[i][index_array_header_2] = ""

    return _array


def sort_by_the_first_column_from_a_to_i(_array):
    """ Отсортировать по первому столбцу от, а до я """
    import locale

    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')

    _shapka = _array[0]
    del _array[0]

    return [_shapka] + sorted(_array, key=lambda x: locale.strxfrm(x[0]))


# Создание графического интерфейса
root = tk.Tk()
root.withdraw()

# Проверка, был ли выбран файл
check_file_open()

# Открытие проводника для выбора файла
file_path = filedialog.askopenfilename()
print("Получили файл")

# Указываем путь к Excel-файлу
# file_path = r'C:\Users\fishzon\Downloads\Pairwise (5).xlsx'

# Первая часть
shapka, array = get_data_and_headers(file_path, 0)
print("Получить данные с первого листа excel таблицы")

array = clear_cells_according_to_the_template(array)
print("Очистили ячейки по шаблону")

# нужно если есть один и тот же параметр со значением да и нет (т.е и да и нет)
#                                       - то покрасить строку в какой-нибудь цвет
array = duplicate_removal(array)
print("Удалили дубликаты")

number_of_runs = int(len(array[0]) / 2)

for i in range(number_of_runs):
    array = split_a_string_with_duplicate_parameters_but_different_values(array)
    print(f"Разбить строку с дубликатами параметров, но разными значениями, сделан прогон {i + 1}")

array = duplicate_removal(array)
print("Еще раз удалить дубликаты")

array = for_column_groups_shift_values_to_empty_cells(array)
print("Для групп столбцов сдвинуть значения в пустые ячейки")

array = sort_by_the_first_column_from_a_to_i(array)
print("Сделана сортировка по первому столбцу от а до я")

save_table_in_new_sheet(file_path, array, 'New Sheet')
print("Сохранить таблицу в новый лист")

open_excel_and_sheet(file_path, 'New Sheet')
print("Открыть эксель, затем открыть лист")
